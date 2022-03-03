import ast
import csv
import datetime
import os
import time
from copy import deepcopy
from datetime import timedelta

import pandas as pd
from py2neo import Graph, Node, Relationship, NodeMatcher

from callback import Callback
from main import main
from sina import getsinadata
import sys

sys.path.append("..")
from train import train
import openpyxl


# 创建节点的同时并创建关系
def createnode(graph, label_name, property_value, rela_name, node_name1):
    node = Node(label_name, name=property_value)
    matcher = NodeMatcher(graph)
    nodelist = list(matcher.match(label_name, name=property_value))
    if len(nodelist) > 0:  # 表示节点存在，不需创建新的节点
        node = nodelist[0]  #
        # 可以直接添加关系
        if rela_name != '':
            rela = Relationship(node_name1, rela_name, node)
            graph.create(rela)
    else:
        # 创建节点
        graph.create(node)
        # 创建关系
        if rela_name != '':
            rela = Relationship(node_name1, rela_name, node)
            graph.create(rela)
    return node


def GetEventNum(graph):
    c = 'match (n:`事件`) return n;'
    data = graph.run(c).data()
    if data:
        data = graph.run(c).to_data_frame()
        json_records = data.to_json(orient="records")
        data = eval(json_records)
        data = pd.DataFrame(data)
        num = []
        for row in data.values:
            p = row[0]['name']
            q = p.replace('事件', '')
            if '评估' not in q and int(q):
                num.append(int(q))
        if num:
            return max(num)
        else:
            return 0
    else:
        return 0


def run():
    while True:
        time_now = time.strftime("%H:%M:%S", time.localtime())  # 刷新
        date_now = time.strftime("%Y/%m/%d", time.localtime())
        print(date_now, time_now)
        time.sleep(0.5)  # 因为以秒定时，所以暂停2秒，使之不会在1秒内执行多次
        if time_now == "23:59:59":  # 此处设置每天定时的时间

            # 首先通过召回获取数据
            a = Callback()
            a.getmsg()
            zh_path = './auto_data/消费品召回.csv'
            zhout_path = './auto_data/result/消费品召回.tsv'
            main(zh_path, zhout_path, _from='召回', needcl=True, key='召回内容', encoding='utf-8', country='中国')
            PutInData(graph, zhout_path)

            # 黑猫平台获取3个月内
            datenow = time.strftime("%Y-%m-%d", time.localtime())
            now = datetime.datetime.now()
            threemothsago = (now - timedelta(days=90)).strftime('%Y-%m-%d')
            getsinadata(threemothsago, datenow, 0)
            zh_path = './auto_data/黑猫投诉.csv'
            zhout_path = './auto_data/result/黑猫投诉.tsv'
            main(zh_path, zhout_path, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
            PutInData(graph, zhout_path)

            Train_File = '../productinfos.csv'
            train(graph, Train_File)


# 数据入库
def PutInData(graph, files):
    harm3 = ['绳索及类似物', '不透气', '填充物', '小零件', '尖角', '锐利边缘', '光滑表面', '粗糙表面', '部件空隙或开口', '机械稳定性', '机械强度', '弹性组件失控',
             '压力空间失控', '移动状态撞击', '旋转部件牵扯', '飞行物体撞击', '移动部件挤压', '爆炸性气体', '爆炸性粉尘', '爆炸性喷雾', '聚合爆炸', '蒸发爆炸', '液体混合爆炸',
             '爆炸性化合物', '固体爆炸性物质', '高/低压', '过热', '漏电', '短路', '过热', '短路', '接触不良', '铁芯发热', '散热不良', '明火', '高温表面', '高温液体',
             '高温气体', '低温表面', '低温液体', '低温气体', '激光辐射', '紫外线辐射', 'X光线辐射', '高频电磁辐射', '低频电磁辐射', '一氧化碳', '一氧化氮', '氯气', '臭氧',
             '氯化氢', '硫化氢', '其它', '砷及其化合物', '镉及其化合物', '铬及其化合物', '铜及其化合物', '汞及其化合物', '镍及其化合物', '铅及其化合物', '其它', '硫酸', '盐酸',
             '氢氧化钠', '其它', '氢氰酸', '氰化钾', '氯化氢', '其它', '甲醛', '乙醛', '丙烯醛', '其它', '蒽类化合物', '菲类化合物', '芘类化合物', '其它',
             'N-杂环化合物', 'S-杂环化合物', 'O-杂环化合物', '其它', '有机氟化物', '有机溴化物', '其它', '大肠杆菌', '沙门氏菌', '副溶血性弧菌', '金黄色葡萄球菌',
             '腊样芽孢肝菌', '其它', '皮肤癣真菌', '着色真菌', '孢子丝菌', '新生隐球菌', '假丝酵母菌', '曲霉', '毛霉', '卡氏肺孢菌', '其它', '甲肝病毒', '甲型流感病毒',
             '轮状病毒', '禽流感病毒', '其它', '尘螨', '蛔虫卵', '绦虫卵', '其它']
    if os.path.exists(files):
        count = len(open(files, 'r', encoding='utf-8').readlines())
        # print(count)
        file = open(files, 'r', encoding='utf-8')
        # last = Getnum(graph)  # 库中共有事件数
        eventnum = GetEventNum(graph)
        counts = deepcopy(eventnum)
        columns = []
        while 1:
            line = file.readline()
            counts = counts + 1
            line = line.replace("\n", "")
            line = line.replace(" ", "")
            p = line.split('\t')
            print(p)
            if line:
                if counts == eventnum + 1:
                    columns = p
                    che = ['砷', '镉', '铬', '铜', '汞', '镍', '铅']
                    chemistry = ['砷及其化合物', '镉及其化合物', '铬及其化合物', '铜及其化合物', '汞及其化合物', '镍及其化合物', '铅及其化合物']
                    for m in range(0, len(columns)):
                        if columns[m] in che:
                            temp = columns[m]
                            columns[m] = temp + '及其化合物'
                            print('标题: 已将', temp, '修改为', columns[m])
                        else:
                            print('标题：', columns[m])
                    for m in range(0, len(columns)):
                        if '\n' in columns[m]:
                            columns[m] = columns[m].replace('\n', '')
                            print('已清楚标题中的换行符')
                        if '' in columns[m]:
                            columns[m] = columns[m].replace(' ', '')
                            print('已清楚标题中的空格')

                    continue
                elif len(columns) != len(p):
                    print('数据不符合要求，已抛弃')
                    continue
                else:
                    date_index = columns.index('日期')
                    title_index = columns.index('事件标题')
                    event_index = columns.index('伤害事件')
                    if JudgeDataExist(graph, p[date_index], p[title_index], p[event_index]):
                        print(p[date_index], p[title_index], '已存在')
                        continue
                    else:
                        print(p[date_index], p[title_index], '不存在')
                        node = createnode(graph, '事件', '事件' + str(GetEventNum(graph) + 1), '', '')
                        print('创建事件', str(GetEventNum(graph) + 1), '中~')
                        # createnode(graph, '数据来源', str(source), '数据来源', node)
                        createnode(graph, '已标注', '0', '已标注', node)
                        severity = []
                        for i in range(1, len(columns)):
                            print(p[i])
                            p[i] = p[i].replace('\"', '')
                            print(p[i])
                            if columns[i] == '消费品一级危害类型' or columns[i] == '消费品二级危害类型' or columns[i] == '消费品危害类型':
                                print('忽略节点:', str(columns[i]))
                                print('进度:', '事件' + str(counts - 1), '%事件' + str(count - 1 + eventnum), '剩余子节点:',
                                      str(len(columns) - i - 1))
                                continue

                            elif columns[i] == '事件来源':
                                if '舆情' in p[i]:
                                    createnode(graph, columns[i], '舆情', columns[i], node)
                                    print('已创建节点:', str(columns[i]), '属性值:', '舆情')
                                elif '投诉' in p[i]:
                                    createnode(graph, columns[i], '投诉', columns[i], node)
                                    print('已创建节点:', str(columns[i]), '属性值:', '投诉')
                                else:
                                    createnode(graph, columns[i], '召回', columns[i], node)
                                    print('已创建节点:', str(columns[i]), '属性值:', '召回')

                            elif columns[i] == '日期':
                                q = p[i].replace('/', '')
                                q = q.replace('\\', '\\')
                                createnode(graph, columns[i], q, columns[i], node)
                                print('已创建节点:', str(columns[i]), '属性值:', q)

                            elif p[i] == '否' or p[i] == '0' or p[i] == '无' or p[i] == '0.0':
                                createnode(graph, columns[i], '0', columns[i], node)
                                print('已创建节点:', str(columns[i]), '属性值:', '0')

                            elif p[i] == '空' or p[i] == '' or p[i] == '-1' or p[i] == [] or p[i] == '[]' or p[
                                i] == '未知' or p[
                                i] == '不详':
                                if columns[i] == '消费品一级类别':
                                    createnode(graph, columns[i], '其他', columns[i], node)
                                    print('已创建节点:', str(columns[i]), '属性值:', '其他')
                                else:
                                    createnode(graph, columns[i], '-1', columns[i], node)
                                    print('已创建节点:', str(columns[i]), '属性值:', '-1')

                            elif p[i] == '有' or p[i] == '是' or p[i] == '1':
                                if columns[i] in harm3:
                                    createnode(graph, '消费品三级危害类型', columns[i], '消费品三级危害类型', node)
                                    print('已创建节点:', '消费品三级危害类型', '属性值:', columns[i])
                                createnode(graph, columns[i], '1', columns[i], node)
                                print('已创建节点:', str(columns[i]), '属性值:', '1')

                            elif ',' in p[i] and columns[i] != '伤害事件':
                                q = p[i].split(',')
                                for var in q:
                                    createnode(graph, columns[i], var, columns[i], node)
                                    print('已创建节点:', str(columns[i]), '属性值:', var)

                            else:
                                createnode(graph, columns[i], p[i], columns[i], node)
                                print('已创建节点:', str(columns[i]), '属性值:', p[i])

                            print('进度:', '事件' + str(counts - 1), '%事件' + str(count - 1 + eventnum), '剩余子节点:',
                                  str(len(columns) - i - 1), '刚创建事件:', str(columns[i]), '属性值:', p[i])

            else:
                break


# 判断数据是否已经存在
def JudgeDataExist(graph, time, title, event):
    # 20210706 【安徽】安徽吉奥塑业有限公司召回部分200ml航空杯 不存在
    q = time.replace('/', '')
    q = q.replace('\\', '\\')
    q = q.replace('\"', '')
    title = title.replace('\"', '')
    title = title.replace('\'', '')
    title = title.replace('\“', '\"')
    title = title.replace('\”', '\"')
    event = event.replace('\"', '')
    event = event.replace('\'', '')
    event = event.replace('\“', '\"')
    event = event.replace('\”', '\"')
    print('debug', q)
    c = 'match (n:`事件`)-[rel]-(d:`日期`{name:"' + q + '"}) with n,rel,d match(n:`事件`)-[rel1]-(t:`事件标题`{name:"' + title + '"}) with n,rel,d,rel1,t match(n:`事件`)-[rel2]-(z:`伤害事件`{name:"' + event + '"}) return n,rel2,z'
    data = graph.run(c).data()

    if data:
        return True
    else:
        return False


# 准备
def DataPreparation(graph):
    # graph.delete_all()
    #
    #PutInData(graph, './auto_data/result/院长基金数据2018_召回通报.txt')  # 数据库名称, 文件路径/文件名, 数据来源[1 召回 2 投诉 3 实际发生新闻]
    #PutInData(graph, './auto_data/result/标准院-国内召回情况-2021年6月有效数据.txt')  # 数据库名称, 文件路径/文件名, 数据来源[1 召回 2 投诉 3 实际发生新闻]
    #PutInData(graph, './auto_data/result/消费品召回1.txt')  # 数据库名称, 文件路径/文件名, 数据来源[1 召回 2 投诉 3 实际发生新闻]
    #PutInData(graph, './auto_data/result/消费品质量安全事件统计2020年5月-6月.txt')  # 数据库名称, 文件路径/文件名, 数据来源[1 召回 2 投诉 3 实际发生新闻]
    # 首先通过召回获取数据
    # a = Callback()
    # a.getmsg()
    zh_path = './auto_data/国内消费品召回.csv'
    zhout_path = './auto_data/result/国内消费品召回.tsv'
    # main(zh_path, zhout_path, _from='召回', needcl=True, key='召回内容', encoding='utf-8', country='中国')
    PutInData(graph, zhout_path)

    zh_path = './auto_data/国外消费品召回.csv'
    zhout_path = './auto_data/result/国外消费品召回.tsv'
    # main(zh_path, zhout_path, _from='召回', needcl=True, key='召回内容', encoding='utf-8', country='中国')
    PutInData(graph, zhout_path)

    # 黑猫平台儿童
    datenow = time.strftime("%Y-%m-%d", time.localtime())
    getsinadata('2017-01-01', datenow, 0, '童装')
    ts_path1 = './auto_data/黑猫投诉童装.csv'
    tsout_path1 = './auto_data/result/黑猫投诉童装.tsv'
    # main(ts_path1, tsout_path1, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
    PutInData(graph, tsout_path1)
    #

    # 黑猫平台水晶泥
    datenow = time.strftime("%Y-%m-%d", time.localtime())
    getsinadata('2017-01-01', datenow, 0, '水晶泥')
    ts_path2 = './auto_data/黑猫投诉水晶泥.csv'
    tsout_path2 = './auto_data/result/黑猫投诉水晶泥.tsv'
    PutInData(graph, tsout_path2)
    # main(ts_path2, tsout_path2, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
    # PutInData(graph, tsout_path2, '投诉')
    # PutInData(graph, '标准院-国内召回情况-2021年6月有效数据.tsv', 1)  # 数据库名称, 文件路径/文件名, 数据来源[1 召回 2 投诉 3 实际发生新闻]
    # PutInData(graph, '标准院-国外召回情况-2021年6月-有效数据.tsv', 1)  # 数据库名称, 文件路径/文件名, 数据来源[1 召回 2 投诉 3 实际发生新闻]
    # PutInData(graph, '消费品召回1.tsv', 1)  # 数据库名称, 文件路径/文件名, 数据来源[1 召回 2 投诉 3 实际发生新闻]
    # PutInData(graph, '消费品质量安全事件统计2020年5月-6月-有效数据.tsv', 1)  # 数据库名称, 文件路径/文件名, 数据来源[1 召回 2 投诉 3 实际发生新闻]
    # print('生成模型中~')
    # Train_File = '../productinfos.csv'
    # train(graph, Train_File)
    # print('生成模型完毕！ 自动化处理数据中')

    # 黑猫平台口罩
    datenow = time.strftime("%Y-%m-%d", time.localtime())
    getsinadata('2017-01-01', datenow, 0, '口罩')
    ts_path3 = './auto_data/黑猫投诉口罩.csv'
    tsout_path3 = './auto_data/result/黑猫投诉口罩.tsv'
    # main(ts_path3, tsout_path3, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
    PutInData(graph, tsout_path3)

    # 黑猫平台儿童文具
    datenow = time.strftime("%Y-%m-%d", time.localtime())
    getsinadata('2017-01-01', datenow, 0, '文具')
    ts_path4 = './auto_data/黑猫投诉文具.csv'
    tsout_path4 = './auto_data/result/黑猫投诉文具.tsv'
    # main(ts_path4, tsout_path4, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
    PutInData(graph, tsout_path4)

    # 黑猫平台家电
    datenow = time.strftime("%Y-%m-%d", time.localtime())
    getsinadata('2017-01-01', datenow, 0, '家电')
    ts_path5 = './auto_data/黑猫投诉家电.csv'
    tsout_path5 = './auto_data/result/黑猫投诉家电.tsv'
    # main(ts_path5, tsout_path5, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
    PutInData(graph, tsout_path5)

    # 黑猫平台手机
    datenow = time.strftime("%Y-%m-%d", time.localtime())
    getsinadata('2017-01-01', datenow, 0, '手机')
    ts_path5 = './auto_data/黑猫投诉手机.csv'
    tsout_path5 = './auto_data/result/黑猫投诉手机.tsv'
    # main(ts_path5, tsout_path5, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
    PutInData(graph, tsout_path5)

    # 黑猫平台电动车
    datenow = time.strftime("%Y-%m-%d", time.localtime())
    getsinadata('2017-01-01', datenow, 0, '电动车')
    ts_path5 = './auto_data/黑猫投诉电动车.csv'
    tsout_path5 = './auto_data/result/黑猫投诉电动车.tsv'
    # main(ts_path5, tsout_path5, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
    PutInData(graph, tsout_path5)

    # 黑猫平台童车
    datenow = time.strftime("%Y-%m-%d", time.localtime())
    getsinadata('2017-01-01', datenow, 0, '童车')
    ts_path5 = './auto_data/黑猫投诉童车.csv'
    tsout_path5 = './auto_data/result/黑猫投诉童车.tsv'
    # main(ts_path5, tsout_path5, _from='投诉', needcl=True, key='投诉内容', encoding='utf-8', country='中国')
    PutInData(graph, tsout_path5)

    
    


# 生成excel文件
def getExcel(graph):
    columns = ['事件号', '事件标题', '日期', '国家', '区域', '消费品一级类别', '消费品名称', '伤害事件', '事件来源', '危害源',  '消费品二级类别', '链接', '伤害类型', '涉及的消费品数量',
               '严重程度', '消费品危害类型', '消费品一级危害类型', '消费品二级危害类型', '消费者年龄', '消费者受教育程度', '消费者职业', '健康状况', '消费品问题部件', '尺寸',
               '绳索及类似物', '不透气', '填充物', '锐利边缘', '部件空隙或开口', '尖角', '光滑表面', '粗糙表面', '小零件', '昼夜', '温度', '湿度', '海拔', '稳定性',
               '腐蚀性', '速度', '腐蚀物', '坎坷', '爬坡', '下坡', '地面摩擦', '斜坡', '楼梯', '灰尘', '静电', '辐射', '警示标识缺失', '高温表面', '高温液体',
               '高温气体', '低温表面', '低温液体', '低温气体', '明火', '高低电压', '过热', '漏电', '短路', '接触不良', '铁芯发热', '散热不良', '稳定性噪音危害',
               '变动性噪音危害', '脉冲性噪音危害', '爆炸性气体', '爆炸性粉尘', '爆炸性喷雾', '聚合爆炸', '蒸发爆炸', '液体混合爆炸', '爆炸性化合物', '固体爆炸性物质',
               '爆炸性气体.1', '旋转部件牵扯', '飞行物体撞击', '移动部件挤压', '机械稳定性', '机械强度', '弹性组件失控', '压力空间失控', '热辐射危害', '激光辐射', '紫外线辐射',
               'X光线辐射', '高频电磁辐射', '低频电磁辐射', '甲醛', '乙醛', '丙烯醛', '其它有毒醛类', '蒽类化合物', '菲类化合物', '芘类化合物', '其它有毒芳香稠环类化合物类',
               'N-杂环化合物', 'S-杂环化合物', 'O-杂环化合物', '其它有毒杂环类化合物', '有机氟化物', '有机氯化物', '有机溴化物', '其它有毒有机氯化物', '一氧化氮', '一氧化碳',
               '氯气', '臭氧', '氯化氢', '硫化氢', '其它有毒气体', '砷及其化合物', '镉及其化合物', '铬及其化合物', '铜及其化合物', '汞及其化合物', '镍及其化合物', '铅及其化合物', '其它有毒重金属及其化合物', '硫酸', '盐酸',
               '氢氧化钠', '其它有毒酸碱类物', '氢氰酸', '氰化钾', '氰化氢', '其它无机氰化物', '尘螨', '蛔虫卵', '绦虫卵', '其它寄生虫危害', '甲肝病毒', '甲型流感病毒',
               '轮状病毒', '禽流感病毒', '其它原生微生物危害', '皮肤癣真菌', '着色真菌', '孢子丝菌', '新生隐球菌', '假丝酵母菌', '曲霉', '毛霉', '卡氏肺孢菌',
               '其它真核细胞微生物危害', '大肠杆菌', '沙门氏菌', '副溶血性弧菌', '金黄色葡萄球菌', '腊样芽孢肝菌', '其它原核细胞微生物危害']
    harm1_0 = {'机械危害': '物理危害', '爆炸危害': '物理危害', '噪声危害': '物理危害', '电气危害': '物理危害', '高/低温物质危害': '物理危害', '辐射危害': '物理危害',
               '警示标识缺失': '物理危害', '无机毒物危害': '化学危害', '有机毒物': '化学危害', '致病微生物危害': '生物危害', '致病生物危害': '生物危害'}
    harm2_1 = {'形状和表面性能危害': '机械危害', '潜在能量危害': '机械危害', '动能危害': '机械危害', '气相爆炸危害': '爆炸危害', '液相爆炸危害': '爆炸危害',
               '固相爆炸危害': '爆炸危害', '稳定性噪音危害': '噪声危害', '变动性噪音危害': '噪声危害', '脉冲性噪音危害': '噪声危害', '触电危害': '电气危害',
               '电气爆炸': '电气危害', '高温物质危害': '高/低温物质危害', '低温物质危害': '高/低温物质危害', '热辐射危害': '辐射危害', '射线辐射危害': '辐射危害',
               '电磁辐射危害': '辐射危害', '警示标识缺失': '警示标识缺失', '有毒气体危害': '无机毒物危害', '有毒重金属及其化合物危害': '无机毒物危害', '有毒酸碱类危害': '无机毒物危害',
               '无机氰化物危害': '无机毒物危害', '有毒荃类化合物': '有机毒物', '有毒芳香稠环类化合物': '有机毒物', '有毒杂环类化合物': '有机毒物', '有毒有机氯化物': '有机毒物',
               '原核细胞微生物危害': '致病微生物危害', '真核细胞微生物危害': '致病微生物危害', '原生微生物危害': '致病微生物危害', '寄生虫危害': '致病生物危害'}
    harm3_2 = {'绳索及类似物': '形状和表面性能危害', '不透气': '形状和表面性能危害', '填充物': '形状和表面性能危害', '小零件': '形状和表面性能危害', '尖角': '形状和表面性能危害',
               '锐利边缘': '形状和表面性能危害', '光滑表面': '形状和表面性能危害', '粗糙表面': '形状和表面性能危害', '部件空隙或开口': '形状和表面性能危害', '机械稳定性': '潜在能量危害',
               '机械强度': '潜在能量危害', '弹性组件失控': '潜在能量危害', '压力空间失控': '潜在能量危害', '移动状态撞击': '动能危害', '旋转部件牵扯': '动能危害',
               '飞行物体撞击': '动能危害', '移动部件挤压': '动能危害', '爆炸性气体': '气相爆炸危害', '爆炸性粉尘': '气相爆炸危害', '爆炸性喷雾': '气相爆炸危害',
               '聚合爆炸': '液相爆炸危害', '蒸发爆炸': '液相爆炸危害', '液体混合爆炸': '液相爆炸危害', '爆炸性化合物': '固相爆炸危害', '固体爆炸性物质': '固相爆炸危害',
               '稳定性噪音危害': '稳定性噪音危害', '变动性噪音危害': '变动性噪音危害', '脉冲性噪音危害': '脉冲性噪音危害', '高/低压': '触电危害', '过热': '电气爆炸',
               '漏电': '触电危害', '短路': '电气爆炸', '接触不良': '电气爆炸', '铁芯发热': '电气爆炸', '散热不良': '电气爆炸', '明火': '高温物质危害',
               '高温表面': '高温物质危害', '高温液体': '高温物质危害', '高温气体': '高温物质危害', '低温表面': '低温物质危害', '低温液体': '低温物质危害',
               '低温气体': '低温物质危害', '热辐射危害': '热辐射危害', '激光辐射': '射线辐射危害', '紫外线辐射': '射线辐射危害', 'X光线辐射': '射线辐射危害',
               '高频电磁辐射': '电磁辐射危害', '低频电磁辐射': '电磁辐射危害', '警示标识缺失': '警示标识缺失', '一氧化碳': '有毒气体危害', '一氧化氮': '有毒气体危害',
               '氯气': '有毒气体危害', '臭氧': '有毒气体危害', '氯化氢': '无机氰化物危害', '硫化氢': '有毒气体危害', '其它': '寄生虫危害',
               '砷及其化合物': '有毒重金属及其化合物危害', '镉及其化合物': '有毒重金属及其化合物危害', '铬及其化合物': '有毒重金属及其化合物危害', '铜及其化合物': '有毒重金属及其化合物危害',
               '汞及其化合物': '有毒重金属及其化合物危害', '镍及其化合物': '有毒重金属及其化合物危害', '铅及其化合物': '有毒重金属及其化合物危害', '硫酸': '有毒酸碱类危害',
               '盐酸': '有毒酸碱类危害', '氢氧化钠': '有毒酸碱类危害', '氢氰酸': '无机氰化物危害', '氰化钾': '无机氰化物危害', '甲醛': '有毒荃类化合物', '乙醛': '有毒荃类化合物',
               '丙烯醛': '有毒荃类化合物', '蒽类化合物': '有毒芳香稠环类化合物', '菲类化合物': '有毒芳香稠环类化合物', '芘类化合物': '有毒芳香稠环类化合物',
               'N-杂环化合物': '有毒杂环类化合物', 'S-杂环化合物': '有毒杂环类化合物', 'O-杂环化合物': '有毒杂环类化合物', '有机氟化物': '有毒有机氯化物',
               '有机氯化物': '有毒有机氯化物', '有机溴化物': '有毒有机氯化物', '大肠杆菌': '原核细胞微生物危害', '沙门氏菌': '原核细胞微生物危害', '副溶血性弧菌': '原核细胞微生物危害',
               '金黄色葡萄球菌': '原核细胞微生物危害', '腊样芽孢肝菌': '原核细胞微生物危害', '皮肤癣真菌': '真核细胞微生物危害', '着色真菌': '真核细胞微生物危害',
               '孢子丝菌': '真核细胞微生物危害', '新生隐球菌': '真核细胞微生物危害', '假丝酵母菌': '真核细胞微生物危害', '曲霉': '真核细胞微生物危害', '毛霉': '真核细胞微生物危害',
               '卡氏肺孢菌': '真核细胞微生物危害', '甲肝病毒': '原生微生物危害', '甲型流感病毒': '原生微生物危害', '轮状病毒': '原生微生物危害', '禽流感病毒': '原生微生物危害',
               '尘螨': '寄生虫危害', '蛔虫卵': '寄生虫危害', '绦虫卵': '寄生虫危害'}
    filename = r'消费品数据.xlsx'
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)
    for col in range(0, len(columns)):
        outws.cell(1, col+1).value = columns[col]

    c = 'match (n:`事件`) return n;'
    data = graph.run(c).data()
    nodes = []
    if data:
        data = graph.run(c).to_data_frame()
        for row in data.values:
            if row[0]['name'] and '评估' not in row[0]['name']:
                nodes.append(row[0]['name'])
    count = 1
    for i in nodes:
        print('生成数据中~ 第', count - 1, '%', len(nodes), '条')
        count = count + 1
        eve = []
        for k in columns:
            eve.append('')
        eve[0] = count - 1
        c1 = "MATCH (n1{name:\"" + str(i) + "\"})- [rel] -> (n2) RETURN n1,type(rel),n2"
        data = graph.run(c1).data()
        if data:
            data = graph.run(c1).to_data_frame()
            json_records = data.to_json(orient="records")
            data = eval(json_records)
            data = pd.DataFrame(data)
            # print(data)
            evename = i
            for row in data.values:
                row[0] = row[0]['name']
                row[2] = row[2]['name']
                # print(row[0], row[2])

                if row[1] == '消费品三级危害类型':
                    harm3 = row[2]
                    harm2 = harm3_2[harm3]
                    harm1 = harm2_1[harm2]
                    harm0 = harm1_0[harm1]
                    index3 = columns.index('危害源')
                    index2 = columns.index('消费品二级危害类型')
                    index1 = columns.index('消费品一级危害类型')
                    index0 = columns.index('消费品危害类型')
                    if eve[index3] and harm3 not in eve[index3]:
                        eve[index3] = eve[index3] + ',' + harm3
                    else:
                        eve[index3] = harm3

                    if eve[index2] and harm2 not in eve[index2]:
                        eve[index2] = eve[index2] + ',' + harm2
                    else:
                        eve[index2] = harm2

                    if eve[index1] and harm1 not in eve[index1]:
                        eve[index1] = eve[index1] + ',' + harm1
                    else:
                        eve[index1] = harm1

                    if eve[index0] and harm0 not in eve[index0]:
                        eve[index0] = eve[index0] + ',' + harm0
                    else:
                        eve[index0] = harm0

                elif row[1] == '伤害类型':
                    # print(row[1], row[2])
                    indexs = columns.index(row[1])
                    if eve[indexs] and row[2] not in eve[indexs]:
                        eve[indexs] = eve[indexs] + ',' + row[2]
                    else:
                        eve[indexs] = row[2]
                    # print(eve[indexs])
                elif row[1] == '链接':
                    indexs = columns.index(row[1])
                    row[2] = row[2].replace('/', '')
                    row[2] = row[2].replace('\\\\', '\\')
                    eve[indexs] = row[2]
                elif row[1] == '小零件':
                    indexs = columns.index(row[1])
                    if eve[indexs] and row[2] not in eve[indexs]:
                        eve[indexs] = eve[indexs] + ',' + row[2]
                    elif str(row[2]) != '-1':
                        eve[indexs] = row[2]

                elif row[1] in columns:
                    indexs = columns.index(row[1])
                    if row[1] != '涉及的消费品数量' and row[1] != '严重程度' and row[1] != '健康状况':
                        if str(row[2]) == '-1':
                            eve[indexs] = '未知'
                        elif str(row[2]) == '0':
                            eve[indexs] = '否'
                        elif str(row[2]) == '1':
                            eve[indexs] = '是'
                        else:
                            eve[indexs] = row[2]
                    else:
                        eve[indexs] = row[2]
            # print(eve[indexhurt])
            indexn = columns.index('严重程度')
            severity = ['', '微弱', '一般', '严重', '非常严重']
            risk = ['', "可接受风险", "低风险", "中风险", "高风险"]
            print(evename)
            if '事件' in evename:
                if eve[indexn]:
                    eve[indexn] = severity[int(eve[indexn])]
                else:
                    eve[indexn] = ''
            elif '评估' in evename:
                if eve[indexn]:
                    eve[indexn] = risk[int(eve[indexn])]
                else:
                    eve[indexn] = ''
            indexn = columns.index('健康状况')
            health = ['', '差', '一般', '良好', '很好']
            if eve[indexn]:
                eve[indexn] = health[int(eve[indexn])]
            else:
                eve[indexn] = ''
            print(eve)
            for p in range(1, len(eve)+1):
                # print(eve[p])
                outws.cell(count, p).value = eve[p-1]
            # lists.append(p)
    outwb.save(filename)


if __name__ == "__main__":
    graph = Graph("http://localhost:7474", user="neo4j", password='123456', name='productinfos')
    # input_path1 = './auto_data/消费品数据-问卷调查-儿童服装.csv'
    # output_path1 = './auto_data/result/消费品数据-问卷调查-儿童服装.tsv'
    # input_path2 = './auto_data/消费品数据-问卷调查-儿童家具.csv'
    # output_path2 = './auto_data/result/消费品数据-问卷调查-儿童家具.tsv'
    # main(input_path1, output_path1, _from='调查问卷', needcl=True, key='伤害事件', encoding='utf-8', country='中国')
    # main(input_path2, output_path2, _from='调查问卷', needcl=True, key='伤害事件', encoding='utf-8', country='中国')
    # PutInData(graph, output_path1, '召回')
    # DataPreparation(graph)
    print('shan chu hzong~')
    c1 = "match (n)-[rel]-(m) delete n,rel,m;"
    c2 = "match (n) delete n"
    graph.run(c1)
    graph.run(c2)
    graph = Graph("http://localhost:7474", user="neo4j", password='123456', name='assessment')
    graph.run(c1)
    graph.run(c2)    # new data
    # graph.delete_all()
    print('yi shan chu')
    graph = Graph("http://localhost:7474", user="neo4j", password='123456', name='productinfos')
    PutInData(graph, './auto_data/result/整合数据2ed.tsv')
    
    # run()
    # getExcel(graph)
    # f = open('./auto_data/result/国内消费品召回.tsv', 'w', encoding='utf-8', newline='')
    # tsv_reader = pd.read_csv('./auto_data/result/黑猫投诉家电.tsv', sep='\t')
    # # # print(tsv_reader.head(0))
    # data = tsv_reader.columns.values
    # # # p = tsv_reader.tolist()
    # for i in data:
    #     print('\'' + i + '\'', end=',')
